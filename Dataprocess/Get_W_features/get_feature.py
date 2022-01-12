from CrossTalk_class import InterCrossTalk
from read_Excel import readExcel
from write_Excel import writeExcel
import openpyxl
from openpyxl.styles import Font

# Inter Cross-talk origin file path
Inter_Cross_talk_file = '../Datasets/Cross-talk.xlsx'
Positive_sheet, Po_rows, Po_cols = readExcel(Inter_Cross_talk_file, 'Inter Cross-talk')
Negative_sheet, Ne_rows, Ne_cols = readExcel(Inter_Cross_talk_file, 'Inter Cross-talk Negative_8126')
# positiveset_file = r'C:\Users\leishen\Desktop\PTM_Cross_talk_Imbalanced_learning\dataset\ptm cross-talk\PositiveDataset.xlsx'
# negativeset_file = r'C:\Users\leishen\Desktop\PTM_Cross_talk_Imbalanced_learning\dataset\ptm cross-talk\NegativeDataset.xlsx'

# feature file path
feature_path = r'C:\Users\leishen\Desktop\master3\Inter Cross-talk\feature\results_feature'
# Str features calculated by bio3d and enm
cij_path = feature_path + r'\cij'  # betweenness, clossness, degree, cluster, diversity, eccentricity, strength,eigen_centrality, page_rank
enm_anm_cancer_cc_path = feature_path + r'\enm\anm\cc'  # 5_per,5_per_20_per, 20_per_50_per, greater_60_per, top3
enm_anm_cancer_prs_path = feature_path + r'\enm\anm\prs'  # effectiveness, prs, sensitivity
enm_anm_cancer_sq_path = feature_path + r'\enm\anm\sq'  # sq
enm_anm_cancer_stiffness_path = feature_path + r'\enm\anm\stiffness'  # stiffness
enm_gnm_cancer_cc_path = feature_path + r'\enm\gnm\cc'  # 5_per,5_per_20_per, 20_per_50_per, greater_60_per, top3
enm_gnm_cancer_eigenvector_path = feature_path + r'\enm\gnm\eigenvector'  # eigenvector_20, eigenvector_all, eigenvector_top3
enm_gnm_cancer_prs_path = feature_path + r'\enm\gnm\prs'  # effectiveness, prs, sensitivity
enm_gnm_cancer_sq_path = feature_path + r'\enm\gnm\sq'  # sq
# Seq features calculated by evol
evol_dirinfo = feature_path + r'\evol\dirinfo'
evol_entropy = feature_path + r'\evol\entropy'
evol_mifc = feature_path + r'\evol\mifc'
evol_mifn = feature_path + r'\evol\mifn'
evol_mutinfo = feature_path + r'\evol\mutinfo'
evol_occupancy = feature_path + r'\evol\occupancy'
evol_omes = feature_path + r'\evol\omes'
evol_sca = feature_path + r'\evol\sca'


def get_feature(sheet, rows, cols, featurefile):
    print(sheet, rows, cols)  # <xlrd.sheet.Sheet object at 0x0000024792E09EC8> 10722 21

    titles = ['Pro1_name', 'UniprotId1', 'UniprotSite1', "Site1Type", 'PdbChain1',
              "PdbSite1", "Pro2_name", "UniprotId2", 'UniprotSite2', 'Site2Type', 'PdbChain2',
              'PdbSite2', 'PPI_flag',
              'betweenness_cij', 'closeness_cij', 'cluster_cij', 'degree_cij', 'eccentricity_cij', 'eigen_centrality_cij',
              'anm_cc_20_per_50_per', 'anm_cc_5_per', 'anm_cc_5_per_20_per', 'anm_cc_greater_60_per',
              'anm_cc_top3', 'anm_effectiveness_all', 'anm_prs_all', 'anm_sensitivity_all', 'anm_sq_all',
              'anm_stiffness', 'gnm_cc_20_per_50_per', 'gnm_cc_5_per', 'gnm_cc_5_per_20_per',
              'gnm_cc_greater_60_per', 'gnm_cc_top3', 'gnm_eigenvectors_20', 'gnm_eigenvectors_all',
              'gnm_eigenvectors_top3', 'gnm_effectiveness_all', 'gnm_prs_all', 'gnm_sensitivity_all',
              'gnm_sq_all', 'evol_dirinfo', 'evol_entropy', 'evol_mifc', 'evol_mifn', 'evol_mutinfo', 'evol_occupancy', 'evol_omes',
              'sca']
    features_name = ['betweenness_cij', 'closeness_cij', 'cluster_cij', 'degree_cij', 'eccentricity_cij', 'eigen_centrality_cij',
                     'anm_cc_20_per_50_per', 'anm_cc_5_per', 'anm_cc_5_per_20_per', 'anm_cc_greater_60_per',
                     'anm_cc_top3', 'anm_effectiveness_all', 'anm_prs_all', 'anm_sensitivity_all', 'anm_sq_all',
                     'anm_stiffness', 'gnm_cc_20_per_50_per', 'gnm_cc_5_per', 'gnm_cc_5_per_20_per',
                     'gnm_cc_greater_60_per', 'gnm_cc_top3', 'gnm_eigenvectors_20', 'gnm_eigenvectors_all',
                     'gnm_eigenvectors_top3', 'gnm_effectiveness_all', 'gnm_prs_all', 'gnm_sensitivity_all',
                     'gnm_sq_all', 'evol_dirinfo', 'evol_entropy', 'evol_mifc', 'evol_mifn', 'evol_mutinfo', 'evol_occupancy', 'evol_omes',
                     'evol_sca']

    for row in range(1, rows + 1):
        # for row in range(1, 5):
        if row == 1:
            Inter_CrossTalk_excel = openpyxl.Workbook()
            sheet1 = Inter_CrossTalk_excel.create_sheet('Sheet', 0)
            for i in range(1, len(titles) + 1):
                sheet1.cell(row, i, value=titles[i - 1])
        else:
            features_all = []
            features_Pro1 = []
            features_Pro2 = []
            print(row)
            print(sheet.row_values(row - 1))
            Pro1_name = str(sheet.row_values(row - 1)[0])
            UniprotId1 = str(sheet.row_values(row - 1)[1])
            UniprotSite1 = str(sheet.row_values(row - 1)[2])
            Site1Type = str(sheet.row_values(row - 1)[3])
            PdbChain1 = str(sheet.row_values(row - 1)[4])
            if sheet.row_values(row - 1)[5] == 'None':
                PdbSite1 = str(sheet.row_values(row - 1)[5])
            else:
                PdbSite1 = int(sheet.row_values(row - 1)[5])
            Pro2_name = str(sheet.row_values(row - 1)[6])
            UniprotId2 = str(sheet.row_values(row - 1)[7])
            UniprotSite2 = str(sheet.row_values(row - 1)[8])
            Site2Type = str(sheet.row_values(row - 1)[9])
            PdbChain2 = str(sheet.row_values(row - 1)[10])
            if sheet.row_values(row - 1)[11] == 'None':
                PdbSite2 = str(sheet.row_values(row - 1)[11])
            else:
                PdbSite2 = int(sheet.row_values(row - 1)[11])
            PPI_flag = str(sheet.row_values(row - 1)[12])
            Inter_CrossTalk = InterCrossTalk(Pro1_name, UniprotId1, UniprotSite1, Site1Type,
                                             PdbChain1, PdbSite1, Pro2_name,
                                             UniprotId2, UniprotSite2, Site2Type, PdbChain2, PdbSite2,
                                             PPI_flag)
            Inter_CrossTalk.pdbchain_feature_cij(cij_path)
            Inter_CrossTalk.pdbchain_feature_anm_cc(enm_anm_cancer_cc_path)
            Inter_CrossTalk.pdbchain_feature_anm_prs(enm_anm_cancer_prs_path)
            Inter_CrossTalk.pdbchain_feature_anm_sq(enm_anm_cancer_sq_path)
            Inter_CrossTalk.pdbchain_feature_anm_stiffness(enm_anm_cancer_stiffness_path)
            Inter_CrossTalk.pdbchain_feature_gnm_cc(enm_gnm_cancer_cc_path)
            Inter_CrossTalk.pdbchain_feature_gnm_eigenvector(enm_gnm_cancer_eigenvector_path)
            Inter_CrossTalk.pdbchain_feature_gnm_prs(enm_gnm_cancer_prs_path)
            Inter_CrossTalk.pdbchain_feature_gnm_sq(enm_gnm_cancer_sq_path)
            # print(Ptm_mutation.UniprotPosition)
            Inter_CrossTalk.uninprot_feature_evol_dirinfo(evol_dirinfo)
            Inter_CrossTalk.uninprot_feature_evol_entropy(evol_entropy)
            Inter_CrossTalk.uninprot_feature_evol_mifc(evol_mifc)
            Inter_CrossTalk.uninprot_feature_evol_mifn(evol_mifn)
            Inter_CrossTalk.uninprot_feature_evol_mutinfo(evol_mutinfo)
            Inter_CrossTalk.uninprot_feature_evol_occupancy(evol_occupancy)
            Inter_CrossTalk.uninprot_feature_evol_omes(evol_omes)
            Inter_CrossTalk.uninprot_feature_evol_sca(evol_sca)
            features_Pro1.extend([Pro1_name, UniprotId1, UniprotSite1, Site1Type,
                                  PdbChain1, PdbSite1])
            features_Pro2.extend([Pro2_name,
                                  UniprotId2, UniprotSite2, Site2Type, PdbChain2, PdbSite2])
            features_all.extend(
                [Pro1_name, UniprotId1, UniprotSite1, Site1Type,
                 PdbChain1, PdbSite1, Pro2_name,
                 UniprotId2, UniprotSite2, Site2Type, PdbChain2, PdbSite2,
                 PPI_flag])
            print('PdbChain_features: ', len(Inter_CrossTalk.uniprot1_pdbchain_feature),
                  len(Inter_CrossTalk.uniprot2_pdbchain_feature))
            print('Uniprot_features: ', len(Inter_CrossTalk.uniprot1_uniprot_feature),
                  len(Inter_CrossTalk.uniprot2_uniprot_feature))
            print('All_features: ', len(Inter_CrossTalk.features))

            print(Inter_CrossTalk.features.keys())
            print(Inter_CrossTalk.uniprot1_pdbchain_feature.keys(), Inter_CrossTalk.uniprot1_uniprot_feature.keys())
            print(Inter_CrossTalk.uniprot2_pdbchain_feature.keys(), Inter_CrossTalk.uniprot2_uniprot_feature.keys())
            for feature_name in features_name:
                features_all.append(Inter_CrossTalk.features[feature_name])
            print(features_all)
            if len(Inter_CrossTalk.uniprot1_pdbchain_feature.keys()) != 28 or len(
                    Inter_CrossTalk.uniprot1_uniprot_feature.keys()) != 8 or len(Inter_CrossTalk.features.keys()) != 36:
                print(row)
                break
            sheet1.append(features_all)

    Inter_CrossTalk_excel.save(featurefile)


get_feature(Positive_sheet, Po_rows, Po_cols, '../Datasets/Cross_talk_feature/Positive_features1112.xls')
get_feature(Negative_sheet, Ne_rows, Ne_cols, '../Datasets/Cross_talk_feature/Negative_features1112.xls')
